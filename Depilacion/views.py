from django.shortcuts import render,redirect,HttpResponse
from django.db import connection

# Create your views here.

def reservas (request):
    id= request.user.id
    print(id)
    if request.user.is_authenticated:
        if request.method == 'POST':
            print(request.POST)
            fecha = request.POST['date']
            hora = request.POST['hora']
            cliente = request.POST['clientes']
            reserva = 'INSERT INTO `depilacion_reservas`( `Fecha_reserva`, `Cliente_id`, `Hora_reserva`) VALUES (\'%s\',%s,\'%s\')'% (fecha,cliente,hora)
            with connection.cursor() as cursor:
                cursor.execute(reserva)
            return redirect('/')
        with connection.cursor() as cursor:
            cliente = 'select id,nombre,apellido from setup_clientes'
            cursor.execute(cliente)
            cliente = cursor.fetchall()
            clientes =[]
            for c in cliente:
                clientes.append(list(c))
            zona = 'select * from depilacion_zonas'
            cursor.execute(zona)
            zona = cursor.fetchall()
            zonas =[]
            for z in zona:
                zonas.append(list(z))
        return render(request, "Reservas/add.html",{'clientes':clientes, 'zonas':zonas})
         #otro caso redireccionamos al login
    return redirect('/login')

def reserva_list (request):
    id = request.user.id
    print(id)
    if request.user.is_authenticated:
        if request.method == 'POST':
            print(request.POST)
            fecha = request.POST['date']
            consulta = """select a.*, CONCAT(Nombre, ' ', Apellido) cliente , d.*  from depilacion_reservas a 
                            inner join setup_clientes b on a.cliente_id = b.id 
                            left join depilacion_confirmacion c on a.id = c.reserva_id
                            left join depilacion_status d on c.status_id = d.id
                            where a.fecha_reserva = \'%s\' order by 2,4 asc""" %fecha
            #print(consulta)
            with connection.cursor() as cursor:
                cursor.execute(consulta)
                consulta = cursor.fetchall()
                consulta_reserva = []
                for z in consulta:
                    consulta_reserva.append(list(z))
                return render(request, "Reservas/list.html",{'consulta': consulta_reserva,'fecha':fecha})
        # otro caso redireccionamos al login
        return render(request, "Reservas/list.html")
    return redirect('/login')

from openpyxl import Workbook

def reserva_xls (request,fecha):
    user_deposit = request.user.id
    #fecha = request.POST['date']
    with connection.cursor() as cursor:
        #print(fecha)
        consulta = """select fecha_reserva,DATE_FORMAT(HORA_RESERVA, '%r') hora , CONCAT(Nombre, ' ', Apellido) cliente from depilacion_reservas a inner join setup_clientes b on a.cliente_id = b.id where a.fecha_reserva = \'{0}\' order by 1,2,3 asc""" .format(fecha)
        cursor.execute(consulta)
        #print(consulta)
        consulta = cursor.fetchall()
        #print (consulta)
        reservas = []
        for r in consulta:
            reservas.append(list(r))
        #print(consulta)
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # titulo
        ws['A1'] = 'Lista de reserva  '
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:E3')
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B4'] = 'Fecha reserva'
        ws['C4'] = 'Hora'
        ws['D4'] = 'Paciente'
        cont = 5
        # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for r in reservas:
            # print(venta)
            ws.cell(row=cont, column=2).value = r[0]
            ws.cell(row=cont, column=3).value = r[1]
            ws.cell(row=cont, column=4).value = r[2]

            cont = cont + 1
        # Establecemos el nombre del archivo
        nombre_archivo = "Listdo de reservas.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    return redirect('/reservas')

def confirmacion_reserva (request,fecha,id):
    if request.user.is_authenticated:
            with connection.cursor() as cursor:
                confirmacion = 'INSERT INTO `depilacion_confirmacion`(`Reserva_id`, `Status_id`) VALUES (%s,1)' %id
                cursor.execute(confirmacion)
                consulta = """select a.*, CONCAT(Nombre, ' ', Apellido) cliente , d.*  from depilacion_reservas a 
                            inner join setup_clientes b on a.cliente_id = b.id 
                            left join depilacion_confirmacion c on a.id = c.reserva_id
                            left join depilacion_status d on c.status_id = d.id
                            where a.fecha_reserva = \'%s\' order by 2,4 asc""" % fecha
                cursor.execute(consulta)
                consulta = cursor.fetchall()
                consulta_reserva = []
                for z in consulta:
                    consulta_reserva.append(list(z))
            return render(request, "Reservas/list.html", {'consulta': consulta_reserva, 'fecha': fecha})

def cancelacion_reserva (request,fecha,id):
    if request.user.is_authenticated:
            with connection.cursor() as cursor:
                confirmacion = 'INSERT INTO `depilacion_confirmacion`(`Reserva_id`, `Status_id`) VALUES (%s,2)' %id
                cursor.execute(confirmacion)
                consulta = """select a.*, CONCAT(Nombre, ' ', Apellido) cliente , d.*  from depilacion_reservas a 
                            inner join setup_clientes b on a.cliente_id = b.id 
                            left join depilacion_confirmacion c on a.id = c.reserva_id
                            left join depilacion_status d on c.status_id = d.id
                            where a.fecha_reserva = \'%s\' order by 2,4 asc""" % fecha
                cursor.execute(consulta)
                consulta = cursor.fetchall()
                consulta_reserva = []
                for z in consulta:
                    consulta_reserva.append(list(z))
            return render(request, "Reservas/list.html", {'consulta': consulta_reserva, 'fecha': fecha})

import pywhatkit

def mensaje(request,id,fecha):
    with connection.cursor() as cursor:
        loc ="SET lc_time_names = 'es_ES'"
        cursor.execute (loc)
        consulta =  """select DATE_FORMAT(fecha_reserva,'%W-%d de %M')fecha_reserva,hora_reserva, CONCAT(Nombre, ' ', Apellido) cliente,telefono , d.*  from depilacion_reservas a 
                            inner join setup_clientes b on a.cliente_id = b.id 
                            left join depilacion_confirmacion c on a.id = c.reserva_id
                            left join depilacion_status d on c.status_id = d.id
                            where a.id = {0} order by 2,4 asc""" .format(id)
        cursor.execute(consulta)
        reserva = cursor.fetchone()
        mensaje = str('Estimada ' + ' ' + reserva[2] + ' ' + 'le recuerdamos su turno de depilacion definitiva en Silvia Orsi estetica avanzada.' + ' '
                      + str(reserva[0]) + ' ' + 'a las '+' '+ str(reserva[1])  +' '+'POR FAVOR CONFIRMAR ! UN BESO!')
        hora = 'SELECT DATE_FORMAT(NOW(),"%H")'
        cursor.execute(hora)
        hora = cursor.fetchone()
        min = 'SELECT DATE_FORMAT(NOW(),"%i")'
        cursor.execute(min)
        min = cursor.fetchone()
        min = int(min[0]) + 1
        print(mensaje)
        print(hora[0])
        print(min)
        pywhatkit.sendwhatmsg_instantly(reserva[3],
                              mensaje,
                              int(hora[0]), True, 5)
        print("Envio Exitoso!")
        consulta = """select a.*, CONCAT(Nombre, ' ', Apellido) cliente , d.*  from depilacion_reservas a 
                                   inner join setup_clientes b on a.cliente_id = b.id 
                                   left join depilacion_confirmacion c on a.id = c.reserva_id
                                   left join depilacion_status d on c.status_id = d.id
                                   where a.fecha_reserva = \'%s\' order by 2,4 asc""" % fecha
        cursor.execute(consulta)
        consulta = cursor.fetchall()
        consulta_reserva = []
        for z in consulta:
            consulta_reserva.append(list(z))
    return render(request, "Reservas/list.html", {'consulta': consulta_reserva, 'fecha': fecha})

def fichas (request):
    if request.user.is_authenticated:
            if request.method == 'POST':
                with connection.cursor() as cursor:
                    pacient ="""select id,CONCAT(Nombre, ' ',Apellido) cliente from setup_clientes"""
                    cursor.execute(pacient)
                    pacient= cursor.fetchall()
                    paciente =[]
                    for p in pacient:
                        paciente.append(list(p))
                    cliente = request.POST['paciente']
                    consulta ="""select a.fecha,foto_tipo, nombre, Apellido, a.id  from  depilacion_depilacion a 
                                inner join setup_clientes d on a.Cliente_id = d.id
                                where a.cliente_id = %s """ %cliente
                    cursor.execute(consulta)
                    fichas =[]
                    consulta = cursor.fetchall()
                   # print(consulta)
                    for f in consulta:
                        fichas.append(list(f))
                    return render(request, "Fichas/list.html", {'paciente': paciente,'fichas':fichas})
            with connection.cursor() as cursor:
                pacient ="select id,CONCAT(Nombre, ' ',Apellido) cliente from setup_clientes"
                cursor.execute(pacient)
                pacient= cursor.fetchall()
                paciente =[]
                for p in pacient:
                    paciente.append(list(p))
    return render(request, "Fichas/list.html",{'paciente':paciente})


